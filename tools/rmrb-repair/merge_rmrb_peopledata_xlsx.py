"""Merge the PeopleData directory with the existing JSONL and annual XLSX text.

PeopleData is the canonical source for date/page/ordinal/title/href.  Existing
JSONL content is preferred when it matches (it may contain manual repairs), and
the annual XLSX is used to fill the remaining records.  The command only writes
staging artifacts; it never replaces the corpus or Elasticsearch data.
"""
from __future__ import annotations

import argparse
import json
import re
import sqlite3
import unicodedata
from collections import Counter
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterator


DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
PAGE_RE = re.compile(r"\d+")
YEAR_RE = re.compile(r"(\d{4})年")
KNOWN_PAGE_CODES = {101: 1, 111: 11, 202: 2, 505: 5, 606: 6, 909: 9, 1212: 12}
TITLE_SEGMENT_RE = re.compile(r"(?:\s{2,}|　+)")
TRAILING_IMAGE_RE = re.compile(r"[（(]\s*图片\s*[）)]\s*$")


def norm(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).casefold()
    return "".join(ch for ch in text if ch.isalnum())


def primary_title(value: Any) -> str:
    """Return the title line before author/byline lines in legacy exports."""
    for line in str(value or "").splitlines():
        if line.strip():
            return line.strip()
    return ""


def title_variants(value: Any) -> set[str]:
    """Return conservative catalog-title variants from a legacy title field."""
    primary = primary_title(value)
    if not primary:
        return set()
    raw_variants = {primary}
    raw_variants.update(part.strip() for part in TITLE_SEGMENT_RE.split(primary) if part.strip())
    for part in tuple(raw_variants):
        without_image = TRAILING_IMAGE_RE.sub("", part).strip()
        if without_image:
            raw_variants.add(without_image)
        if "——" in without_image:
            heading = without_image.split("——", 1)[0].strip()
            if len(norm(heading)) >= 4:
                raw_variants.add(heading)
    return {normalized for part in raw_variants if (normalized := norm(part))}


def character_bag(value: Any) -> str:
    return "".join(sorted(norm(value)))


def has_content(row: dict[str, Any] | None) -> bool:
    return bool(row and str(row.get("content") or "").strip())


def source_signature(row: dict[str, Any]) -> tuple[int | None, str, str]:
    """Identify byte-equivalent source records that are safe to collapse."""
    return (
        page_number(row.get("page")),
        norm(row.get("title")),
        str(row.get("content") or "").strip(),
    )


def title_indexes(
    rows: list[dict[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    result: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        values = {norm(row.get("title")), *title_variants(row.get("title"))}
        primary_key = norm(primary_title(row.get("title")))
        if len(primary_key) >= 8:
            values.add(f"bag:{character_bag(primary_key)}")
        for key in values:
            if key:
                result.setdefault(key, []).append(row)
    return result


def indexed_candidates(
    canonical: dict[str, Any],
    rows: list[dict[str, Any]],
    used: set[int],
    index: dict[str, list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    """Use O(1) exact lookups and scan the page only for fuzzy fallbacks."""
    key = norm(canonical.get("title"))
    exact: list[dict[str, Any]] = []
    seen: set[int] = set()
    lookup_keys = [key]
    if len(key) >= 8:
        lookup_keys.append(f"bag:{character_bag(key)}")
    for lookup_key in lookup_keys:
        for row in index.get(lookup_key, []):
            identity = id(row)
            if identity not in used and identity not in seen:
                exact.append(row)
                seen.add(identity)
    if exact:
        return exact
    return [row for row in rows if id(row) not in used]


def page_number(value: Any) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        parsed = value
    elif isinstance(value, float) and value.is_integer():
        parsed = int(value)
    else:
        match = PAGE_RE.search(str(value or ""))
        parsed = int(match.group()) if match else None
    return KNOWN_PAGE_CODES.get(parsed, parsed)


def xlsx_paths(root: Path) -> list[tuple[int, Path]]:
    result: list[tuple[int, Path]] = []
    for path in root.glob("*.xlsx"):
        match = YEAR_RE.search(path.name)
        if match:
            result.append((int(match.group(1)), path))
    return sorted(result)


def iter_jsonl_by_date(path: Path) -> Iterator[tuple[str, list[dict[str, Any]]]]:
    current = ""
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, 1):
            row = json.loads(line)
            date = str(row.get("date") or "")[:10]
            if not DATE_RE.fullmatch(date):
                continue
            if current and date < current:
                raise RuntimeError(f"JSONL is not date ordered at line {line_number}: {date}")
            if current and date != current:
                yield current, rows
                rows = []
            current = date
            rows.append(row)
    if current:
        yield current, rows


def iter_xlsx_by_date(root: Path) -> Iterator[tuple[str, list[dict[str, Any]]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required") from exc
    current = ""
    rows: list[dict[str, Any]] = []
    for year, path in xlsx_paths(root):
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            headers = tuple(str(v or "").strip() for v in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
            if headers[:5] != ("年份", "日期", "报纸版次", "标题", "文本内容"):
                raise RuntimeError(f"Unexpected headers in {path}: {headers}")
            for row_number, raw in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                year_value, date_raw, page_raw, title_raw, content_raw = (tuple(raw) + (None,) * 5)[:5]
                # The annual exports use both ``YYYY-MM-DD`` and
                # ``YYYY/MM/DD`` depending on the year.
                date = str(date_raw or "")[:10].replace("/", "-")
                if year_value is None or not DATE_RE.fullmatch(date):
                    continue
                if not date.startswith(f"{year:04d}-"):
                    raise RuntimeError(f"Out-of-year row {path}:{row_number}: {date}")
                if current and date < current:
                    raise RuntimeError(f"XLSX is not date ordered at {path}:{row_number}: {date}")
                if current and date != current:
                    yield current, rows
                    rows = []
                current = date
                rows.append({
                    "date": date,
                    "page": page_number(page_raw),
                    "title": str(title_raw or "").strip(),
                    "content": str(content_raw or "").strip(),
                    "source": f"{path}:{row_number}",
                })
        finally:
            workbook.close()
    if current:
        yield current, rows


def choose(canonical: dict[str, Any], candidates: list[dict[str, Any]]) -> tuple[dict[str, Any] | None, str]:
    if not candidates:
        return None, "none"
    ctitle = norm(canonical["title"])
    same_page = [row for row in candidates if row.get("page") == canonical.get("page")]
    pool = same_page or candidates
    exact = [row for row in pool if ctitle and norm(row.get("title")) == ctitle]
    # The annual exports sometimes repeat the same article block.  Collapse
    # exact duplicates by title + body prefix, retaining the fullest body.
    collapsed: dict[tuple[str, str], dict[str, Any]] = {}
    for row in exact:
        key = (norm(row.get("title")), norm(row.get("content"))[:400])
        prior = collapsed.get(key)
        if prior is None or len(str(row.get("content") or "")) > len(str(prior.get("content") or "")):
            collapsed[key] = row
    exact = list(collapsed.values())
    if len(exact) == 1:
        return exact[0], "exact_title"
    # The legacy JSONL commonly stores ``title\nauthor`` in its title field,
    # while PeopleData exposes only the catalog title.  Matching the first
    # non-empty line is exact (not fuzzy) and is safe only when it identifies a
    # single article body on the same date and page.
    primary = [row for row in pool if ctitle and norm(primary_title(row.get("title"))) == ctitle]
    collapsed = {}
    for row in primary:
        key = (norm(row.get("title")), norm(row.get("content"))[:400])
        prior = collapsed.get(key)
        if prior is None or len(str(row.get("content") or "")) > len(str(prior.get("content") or "")):
            collapsed[key] = row
    primary = list(collapsed.values())
    if len(primary) == 1:
        return primary[0], "exact_primary_title"
    variants = [row for row in pool if ctitle and ctitle in title_variants(row.get("title"))]
    collapsed = {}
    for row in variants:
        key = (norm(row.get("title")), norm(row.get("content"))[:400])
        prior = collapsed.get(key)
        if prior is None or len(str(row.get("content") or "")) > len(str(prior.get("content") or "")):
            collapsed[key] = row
    variants = list(collapsed.values())
    if len(variants) == 1:
        return variants[0], "exact_title_variant"
    if len(ctitle) >= 8:
        reordered = [
            row
            for row in pool
            if character_bag(primary_title(row.get("title"))) == character_bag(ctitle)
        ]
        collapsed = {}
        for row in reordered:
            key = (norm(row.get("title")), norm(row.get("content"))[:400])
            prior = collapsed.get(key)
            if prior is None or len(str(row.get("content") or "")) > len(str(prior.get("content") or "")):
                collapsed[key] = row
        reordered = list(collapsed.values())
        if len(reordered) == 1:
            return reordered[0], "exact_title_characters"
    scored: list[tuple[float, dict[str, Any]]] = []
    for row in pool:
        full_title = norm(row.get("title"))
        title = norm(primary_title(row.get("title"))) or full_title
        if not title or not ctitle:
            continue
        if len(ctitle) >= 6 and (
            ctitle in title or title in ctitle or ctitle in full_title or full_title in ctitle
        ):
            score = 0.99
        else:
            score = max(
                SequenceMatcher(None, ctitle, title, autojunk=False).ratio(),
                SequenceMatcher(None, ctitle, full_title, autojunk=False).ratio(),
                *(SequenceMatcher(None, ctitle, variant, autojunk=False).ratio()
                  for variant in title_variants(row.get("title"))),
            )
        threshold = 0.90 if min(len(ctitle), len(title)) >= 12 else 0.96
        if score >= threshold:
            scored.append((score, row))
    scored.sort(key=lambda item: item[0], reverse=True)
    if scored and (len(scored) == 1 or scored[0][0] - scored[1][0] >= 0.03):
        return scored[0][1], "fuzzy_title"
    return None, "ambiguous"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--directory", type=Path, required=True)
    parser.add_argument("--jsonl", type=Path, required=True)
    parser.add_argument("--xlsx-root", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--unmatched", type=Path, required=True)
    parser.add_argument(
        "--jsonl-unaligned",
        "--jsonl-orphans",
        dest="jsonl_unaligned",
        type=Path,
        help="Audit rows preserved from JSONL without a confident PeopleData alignment",
    )
    parser.add_argument("--report", type=Path, required=True)
    args = parser.parse_args()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.unmatched.parent.mkdir(parents=True, exist_ok=True)
    jsonl_unaligned_path = args.jsonl_unaligned or args.output.with_name(
        f"{args.output.stem}-jsonl-unaligned.jsonl"
    )
    jsonl_unaligned_path.parent.mkdir(parents=True, exist_ok=True)

    directory = sqlite3.connect(f"file:{args.directory}?mode=ro", uri=True)
    dates = [row[0] for row in directory.execute("SELECT issue_date FROM issues WHERE result_count > 0 ORDER BY issue_date")]
    json_iter = iter_jsonl_by_date(args.jsonl)
    xlsx_iter = iter_xlsx_by_date(args.xlsx_root)
    json_date, json_rows = next(json_iter, (None, []))
    xlsx_date, xlsx_rows = next(xlsx_iter, (None, []))
    counters = Counter()
    with (
        args.output.open("w", encoding="utf-8") as out,
        args.unmatched.open("w", encoding="utf-8") as missing,
        jsonl_unaligned_path.open("w", encoding="utf-8") as jsonl_unaligned,
    ):
        def preserve_jsonl_rows(
            rows: list[dict[str, Any]], used: set[int], reason: str, next_ordinal: int
        ) -> int:
            selected_signatures = {
                source_signature(row)
                for row in rows
                if id(row) in used and has_content(row)
            }
            for row in rows:
                if not has_content(row):
                    counters["jsonlEmptyContentRows"] += 1
                    continue
                counters["jsonlContentRows"] += 1
                if id(row) in used:
                    counters["jsonlMatchedContentRows"] += 1
                    continue
                if source_signature(row) in selected_signatures:
                    counters["jsonlDuplicateRowsCollapsed"] += 1
                    continue
                page = page_number(row.get("page"))
                preserved = {
                    "date": str(row.get("date") or "")[:10],
                    "page": int(page or 0),
                    "ordinal": next_ordinal,
                    "title": str(row.get("title") or "").strip(),
                    "href": None,
                    "content": str(row.get("content") or "").strip(),
                    "contentSource": "jsonl",
                    "matchMethod": "jsonl_source_preserved",
                    "sourceOnly": True,
                }
                out.write(json.dumps(preserved, ensure_ascii=False) + "\n")
                jsonl_unaligned.write(
                    json.dumps(
                        {
                            **row,
                            "alignmentReason": reason,
                            "preservedOrdinal": next_ordinal,
                        },
                        ensure_ascii=False,
                    )
                    + "\n"
                )
                next_ordinal += 1
                counters["jsonlUnalignedContentRows"] += 1
                counters["jsonlPreservedSourceOnlyRows"] += 1
            return next_ordinal

        for index, date in enumerate(dates, 1):
            while json_date is not None and json_date < date:
                counters["jsonOnlyRows"] += len(json_rows)
                preserve_jsonl_rows(json_rows, set(), "date_not_in_peopledata_directory", 0)
                json_date, json_rows = next(json_iter, (None, []))
            while xlsx_date is not None and xlsx_date < date:
                counters["xlsxOnlyRows"] += len(xlsx_rows)
                xlsx_date, xlsx_rows = next(xlsx_iter, (None, []))
            local = json_rows if json_date == date else []
            xrows = xlsx_rows if xlsx_date == date else []
            local_by_page: dict[int | None, list[dict[str, Any]]] = {}
            xlsx_by_page: dict[int | None, list[dict[str, Any]]] = {}
            for row in local:
                local_by_page.setdefault(page_number(row.get("page")), []).append(row)
            for row in xrows:
                xlsx_by_page.setdefault(row.get("page"), []).append(row)
            local_indexes = {page: title_indexes(rows) for page, rows in local_by_page.items()}
            xlsx_indexes = {page: title_indexes(rows) for page, rows in xlsx_by_page.items()}
            used_local: set[int] = set()
            used_xlsx: set[int] = set()
            directory_rows = list(directory.execute(
                "SELECT ordinal, page_number, title, href FROM articles WHERE issue_date = ? ORDER BY ordinal", (date,)
            ))
            for ordinal, page, title, href in directory_rows:
                canonical = {"date": date, "page": page, "ordinal": ordinal, "title": title, "href": href}
                local_page_rows = local_by_page.get(page, [])
                local_candidates = indexed_candidates(
                    canonical,
                    local_page_rows,
                    used_local,
                    local_indexes.get(page, {}),
                )
                local_match, local_method = choose(canonical, local_candidates)
                # Older annual exports omit the page column entirely; in that
                # case the date/title still provide a valid match.
                xlsx_match: dict[str, Any] | None = None
                xlsx_method = "not_needed"
                if not has_content(local_match):
                    xlsx_candidates = list(xlsx_by_page.get(page, []))
                    if not xlsx_candidates:
                        xlsx_candidates = list(xlsx_by_page.get(None, []))
                        xlsx_index = xlsx_indexes.get(None, {})
                    else:
                        xlsx_index = xlsx_indexes.get(page, {})
                    xlsx_candidates = indexed_candidates(
                        canonical, xlsx_candidates, used_xlsx, xlsx_index
                    )
                    xlsx_match, xlsx_method = choose(canonical, xlsx_candidates)
                if local_match is not None:
                    used_local.add(id(local_match))
                if xlsx_match is not None:
                    used_xlsx.add(id(xlsx_match))
                # JSONL is the trusted baseline when it has a body.  An empty
                # JSONL row must not mask a non-empty annual-XLSX fallback.
                selected = (
                    local_match if has_content(local_match)
                    else xlsx_match if has_content(xlsx_match)
                    else local_match or xlsx_match
                )
                source = (
                    "jsonl" if selected is local_match and selected is not None
                    else "xlsx" if selected is xlsx_match and selected is not None
                    else None
                )
                method = (
                    local_method if source == "jsonl"
                    else xlsx_method if source == "xlsx"
                    else "none"
                )
                if selected:
                    content = str(selected.get("content") or "").strip()
                    if content:
                        counters[f"matched_{source}_{method}"] += 1
                    else:
                        counters["emptyMatchedContent"] += 1
                else:
                    counters["missingContent"] += 1
                    missing.write(json.dumps(canonical, ensure_ascii=False) + "\n")
                    content = ""
                record = {**canonical, "content": content, "contentSource": source, "matchMethod": method}
                out.write(json.dumps(record, ensure_ascii=False) + "\n")
            counters["directoryDates"] += 1
            if json_date == date:
                next_ordinal = max((int(row[0]) for row in directory_rows), default=-1) + 1
                preserve_jsonl_rows(
                    json_rows, used_local, "unmatched_on_peopledata_date", next_ordinal
                )
                json_date, json_rows = next(json_iter, (None, []))
            if xlsx_date == date:
                xlsx_date, xlsx_rows = next(xlsx_iter, (None, []))
            if index % 100 == 0:
                print(f"processed {index}/{len(dates)} dates", flush=True)
        while json_date is not None:
            counters["jsonOnlyRows"] += len(json_rows)
            preserve_jsonl_rows(json_rows, set(), "date_after_peopledata_scope", 0)
            json_date, json_rows = next(json_iter, (None, []))
        while xlsx_date is not None:
            counters["xlsxOnlyRows"] += len(xlsx_rows)
            xlsx_date, xlsx_rows = next(xlsx_iter, (None, []))
    accounted_jsonl = (
        counters["jsonlMatchedContentRows"]
        + counters["jsonlDuplicateRowsCollapsed"]
        + counters["jsonlPreservedSourceOnlyRows"]
    )
    counters["jsonlAccountedContentRows"] = accounted_jsonl
    counters["jsonlOrphanedContentRows"] = counters["jsonlContentRows"] - accounted_jsonl
    safe_to_publish = counters["jsonlOrphanedContentRows"] == 0
    report = {
        "scope": {"start": dates[0] if dates else None, "end": dates[-1] if dates else None},
        "directory": str(args.directory.resolve()),
        "jsonl": str(args.jsonl.resolve()),
        "xlsxRoot": str(args.xlsx_root.resolve()),
        "output": str(args.output.resolve()),
        "unmatched": str(args.unmatched.resolve()),
        "jsonlUnaligned": str(jsonl_unaligned_path.resolve()),
        "safeToPublish": safe_to_publish,
        "counters": dict(sorted(counters.items())),
    }
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False))
    if not safe_to_publish:
        raise SystemExit(
            "Refusing a publishable merge: "
            f"{counters['jsonlOrphanedContentRows']} non-empty JSONL rows are unrepresented; "
            f"inspect {jsonl_unaligned_path}"
        )


if __name__ == "__main__":
    main()
